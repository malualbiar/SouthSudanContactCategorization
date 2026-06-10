import pandas as pd
from openpyxl import Workbook

contacts = "South_Sudan_Phone_Numbers.xlsx"
Categorised_Contacts = "South_Sudan_Phone_Numbers_Categorized.xlsx"

#categories
output = {
    "MTN Corporate 1": [],
    "MTN Corporate 2": [],
    "MTN Corporate 3": [],
    "MTN Numbers": [],
    "Zain Numbers": [],
    "Digital Numbers": []
}

df = pd.read_excel(contacts)

# flatten everything 
all_numbers = df.astype(str).values.flatten()

#classification logic
for value in all_numbers:

    if pd.isna(value):
        continue

    # Remove spaces and non-digit characters
    number = "".join(filter(str.isdigit, str(value)))

    if not number:
        continue

    # Corporate MTN Groups
    if number.startswith("211922903"):
        output["MTN Corporate 1"].append(number)

    elif number.startswith("211922904"):
        output["MTN Corporate 2"].append(number)

    elif number.startswith("211922905"):
        output["MTN Corporate 3"].append(number)

    # General MTN Numbers
    elif number.startswith("21192"):
        output["MTN Numbers"].append(number)

    # Zain
    elif number.startswith("21191"):
        output["Zain Numbers"].append(number)

    # Digital
    elif number.startswith("21198"):
        output["Digital Numbers"].append(number)

#export to excel
wb = Workbook()

# Remove default worksheet
default_sheet = wb.active
wb.remove(default_sheet)

for sheet_name, numbers in output.items():

    ws = wb.create_sheet(title=sheet_name)

    # Header Row
    ws["A1"] = "Phone Number"

    # Write Numbers
    for row, phone_number in enumerate(numbers, start=2):
        ws.cell(row=row, column=1, value=phone_number)


wb.save(Categorised_Contacts)